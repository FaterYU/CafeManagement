import string
from tokenize import String
import openpyxl
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
import pymysql
import sys
import time
import json

def parser_merged_cell(sheet: Worksheet, row, col):
    """
    检查是否为合并单元格并获取对应行列单元格的值。
    如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
    :param sheet: 当前工作表对象
    :param row: 需要获取的单元格所在行
    :param col: 需要获取的单元格所在列
    :return: 
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row,
                                  column=merged_range.min_col)
                break
    return cell


def get_cell_value(path):
    wb = openpyxl.load_workbook(path)
    # 获得所有 sheet 的名称()
    name = wb.sheetnames[0]
    
    sheet = []
    my_sheet = wb[name]
# 获得 sheet 名
    print(my_sheet.title)
    # iter_rows() 方法获得多个单元格
    print(my_sheet.max_row)
    print(my_sheet.max_column)
    for row_index in range(2, my_sheet.max_row+1):
        row = []
        # print(row_index)
        for col_index in range(1, my_sheet.max_column+1):
            
            cell = my_sheet.cell(row_index, col_index)
            # print(cell)
            row.append(str(cell.value))
            # print(row)
        sheet.append(row)
    # print(table[0])
    return [sheet]


def database(table):
    # 打开数据库连接
    db = pymysql.connect(host='localhost',
                         user='cfm',
                         password='cfm123456',
                         database='cfmdb')

    cursor = db.cursor()

    # 使用 execute() 方法执行 SQL，如果表存在则删除
    cursor.execute("DROP TABLE IF EXISTS `goods`")

    # 使用预处理语句创建表
    sql = """CREATE TABLE `goods` (
            `id`  INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
            `createdAt`  TEXT NOT NULL,
            `updatedAt`  TEXT NOT NULL,
            `name`  TEXT NOT NULL,
            `class`  TEXT NOT NULL,
            `ask`  JSON,
            `sale`  BOOLEAN,
            `image_name`  TEXT NOT NULL,
            `image`  TEXT,
            `price`  FLOAT,
            `describe`  TEXT )"""

    cursor.execute(sql)
    # 构造数据
    allsql = []
    label = ['`name`', '`class`', '`ask`', '`sale`', '`image_name`',
             '`image`', '`price`', '`describe`']
    for row in range(1, len(table[0])):
        sql = ["INSERT INTO goods(`id`,`createdAt`,`updatedAt`,"]
        for col in range(len(label)):
            # if(table[0][row][col]!='None'):
            sql.append(label[col])
            if(col != len(label)-1):
                sql.append(",")
        sql.append(") VALUES (")
        sql.append("'"+str(row)+"',")
        sql.append(time.strftime("'%Y-%m-%d %H:%M:%S'", time.localtime())+',')
        sql.append(time.strftime("'%Y-%m-%d %H:%M:%S'", time.localtime())+',')

        sql.append("'"+str(table[0][row][1])+"',")
        sql.append("'"+str(table[0][row][0])+"',")
        temp=[]
        if(table[0][row][2]):
            temp.append('Hot')
        if(table[0][row][3]):
            temp.append('Cold')
        ask = {'temp':temp}
        d_ask = pymysql.converters.escape_string(json.dumps(ask))
        sql.append("'"+d_ask+"',")
        sql.append("'"+str(1)+"',")
        sql.append("'"+str(table[0][row][5])+"',")
        sql.append("'"+str(0)+"',")
        sql.append("'"+str(table[0][row][4])+"',")
        sql.append("'"+str('unknown')+"'")
        
        sql.append(")")
        # print(''.join(''.join(sql).split('\n'))+'\n')
        allsql.append(''.join(''.join(sql).split('\n')))
    # print(allsql[0])
    
    for sql in allsql:
        try:
            # 执行sql语句
            cursor.execute(sql)
            # 提交到数据库执行
            db.commit()
        except:
            # 如果发生错误则回滚
            db.rollback()
    
    # # 检查数据库
    # sql="select * from origindata"
    # cursor.execute(sql)
    # results = cursor.fetchall()
    # for row in results:
    #     print(row)

    # 关闭数据库连接
    db.close()
    print("数据库连接已关闭")


# main
table = get_cell_value(str(sys.argv[1]))  # 文件路径
# table=get_cell_value('E:\workspace\Project\CafeManagement\sh\goods_msg.xlsx')
# print(table)
database(table)
